from openpyxl import Workbook, load_workbook
from datetime import date
import pandas as pd
import argparse
import warnings
import logging
import sys
import os

# Was written by Ellie, Hodjayev - for problems please contact.

# Create logger
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    handlers=[
                        logging.FileHandler(filename="info.log"),
                        logging.StreamHandler(sys.stdout)
                    ])
logger = logging.getLogger()


def inventory():
    """This will allow interaction within MMEX Inventory.xlsx."""

    # Save date and computer date (reversed)
    today = date.today()
    computer_date = today.strftime("%m-%d-%Y")

    # Ignore warnings - set up for xlsx files with data validation
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

    # Load xlsx - pandas
    file = "S:\\Groups\\FV\\Memory\\everything_else\\MMEX\\InventoryFiles\\MMEX Inventory.xlsx"
    df = pd.ExcelFile(file).sheet_names

    # Load xlsx - openpyxl
    wb = load_workbook(filename=file)

    # Filter sheets
    counter = 0
    sheets = []
    for sheet in df:
        if sheet == "EXTRA" or sheet == "Inventory Rules" or sheet == "Removed lines" or sheet == "EOL_Hynix_SODIMM" \
                or sheet == "EV" or sheet == "LPDDR4" or sheet == "LP4":
            pass
        else:
            counter += 1
            sheets.append(sheet)

    # Added arguments to take
    parser = argparse.ArgumentParser(description="Will allow interaction within Inventory")
    parser.add_argument("num", help="What memory are you looking for? min of 2 letters are "
                                    "sufficient.")
    parser.add_argument("-m", "--subtract", type=int, metavar='', help="Will add to mmex and subtract from cabinet")
    parser.add_argument("-c", "--add", type=int, metavar='', help="Will add to cabinet and subtract from mmex")
    args = parser.parse_args()

    # Loop through sheets
    sheet_counter = 0
    for i in sheets:
        if sheet_counter == len(sheets):
            break
        else:
            # Read xlsx and current sheet
            df = pd.read_excel(f"{file}", f"{sheets[sheet_counter]}")
            ws = wb[f"{sheets[sheet_counter]}"]

            # Look for 'MMEX' column
            counter = 0
            for col in df.columns:
                if col == 'VDR':
                    break
                else:
                    counter += 1
            vdr_column = counter + 1

            # Look for 'Cabinet Qty' column
            counter = 0
            for col in df.columns:
                if col == 'Cabinet Qty':
                    break
                else:
                    counter += 1
            cabinet_column = counter + 1

            # Look for IDC S/N column
            counter = 0
            for col in df.columns:
                if col == 'IDC S/N':
                    break
                else:
                    counter += 1
            idc_column = counter

            # Will allow conversion of letters to number(s)
            characters = 'abcdefghijklmnopqrstuvwxyz'

            # Converted numbers to letters - 1 = A , 2 = B
            header = characters[idc_column]

            # Compare and keep matching columns
            a = df.columns
            b = ['IDC S/N', 'ECC', 'Cabinet Qty', 'MMEX', 'VDR']
            keep_columns = [x for x in a if x in b]

            # Maximum width on output
            pd.set_option('display.max_columns', None)
            pd.set_option('display.width', None)
            pd.set_option('display.max_colwidth', None)

            # Search within IDC S/N for argument
            try:
                df = df.loc[df['IDC S/N'].str.lower().str.contains(args.num.lower(), na=False), keep_columns]
                df.reset_index(drop=True, inplace=True)
            except AttributeError as e:
                pass
            finally:
                pass

            counter = 0
            if df.empty:
                pass
            else:
                serial = df.iloc[0]['IDC S/N']  # Serial number only (Memory)
                # Search for memory within 'IDC S/N'
                for cell in ws[header]:
                    if cell.value == serial:
                        row = cell.row
                        break
                    else:
                        pass

            # Enable user to edit 'Cabinet Qty' or 'VDR'
            if args.add:
                if df.empty:
                    pass
                else:
                    # Check whether calculation approves
                    check = ws.cell(row=row, column=vdr_column).value

                    if check - args.add < 0:
                        logger.error(f"\n\n\n{df}\nYou cannot do that.\n"
                                     f"While available quantity on VDR is "
                                     f"{ws.cell(row=row, column=vdr_column).value}\n"
                                     f"You are trying to subtract it by {args.add}\n")
                        exit()

                    # Convert from float to int
                    try:
                        df['MMEX'] = df['MMEX'].fillna('nan').astype(int),
                        df['VDR'] = df['VDR'].fillna('nan').astype(int)
                    except KeyError:
                        pass
                    finally:
                        pass

                    # Log user and changes
                    logger.info(f"Being edited by - {os.getlogin()}")
                    logger.info(f"The following changes are being made in sheet - {sheets[sheet_counter]}\n"
                                f"Memory - {serial}\n"
                                f"Cabinet Qty - {ws.cell(row=row, column=cabinet_column).value}\n"
                                f"VDR - {ws.cell(row=row, column=vdr_column).value}\n")

                    # Make changes to 'Cabinet Qty/MMEX'
                    ws.cell(row=row, column=cabinet_column).value += args.add
                    ws.cell(row=row, column=vdr_column).value -= args.add

                    # Save changes
                    wb.save(file)
                    logger.info(f"The following changes have been made\n"
                                f"Memory - {serial}\n"
                                f"Cabinet Qty - {ws.cell(row=row, column=cabinet_column).value}\n"
                                f"VDR - {ws.cell(row=row, column=vdr_column).value}\n")

            elif args.subtract:
                if df.empty:
                    pass
                else:
                    # Check whether calculation approves
                    check = ws.cell(row=row, column=cabinet_column).value
                    if check - args.subtract < 0:
                        logger.error(f"\n\n\n{df}\nYou cannot do that.\n"
                                     f"While available quantity on 'Cabinet Qty' is {args.add}\n"
                                     f"You are trying to subtract it by {args.subtract}\n\n")
                        exit()

                    # Convert from float to int
                    try:
                        df['MMEX'] = df['MMEX'].fillna('nan').astype(int)
                        df['VDR'] = df['VDR'].fillna('nan').astype(int)
                    except KeyError:
                        pass
                    finally:
                        pass

                    # Log user and changes
                    logger.info(f"Being edited by - {os.getlogin()}")
                    logger.info(f"The following changes are being made in sheet - {sheets[sheet_counter]}\n\n"
                                f"Memory - {serial}\n"
                                f"Cabinet Qty - {ws.cell(row=row, column=cabinet_column).value}\n"
                                f"VDR - {ws.cell(row=row, column=vdr_column).value}\n")

                    # Make Changes to 'Cabinet Qty/MMEX'
                    ws.cell(row=row, column=cabinet_column).value -= args.subtract
                    ws.cell(row=row, column=mmex_column).value += args.subtract

                    # Save changes
                    wb.save(file)
                    logger.info(f'The following changes have been made in\n'
                                f"Memory - {serial}\n"
                                f"Cabinet Qty - {ws.cell(row=row, column=cabinet_column).value}\n"
                                f"VDR - {ws.cell(row=row, column=vdr_column).value}\n")

            # Will prevent empty dataframes when looping from sheets
            if df.empty:
                sheet_counter += 1
            else:
                # Prevents overwriting over a second sheet
                if args.subtract or args.add:
                    break
                else:
                    print(f"\n{sheets[sheet_counter]}\n"
                          f"Memory - {serial}\n"
                          f"Cabinet Qty - {ws.cell(row=row, column=cabinet_column).value}\n"
                          f"VDR - {ws.cell(row=row, column=vdr_column).value}\n")
                    sheet_counter += 1


if __name__ == "__main__":
    inventory()
